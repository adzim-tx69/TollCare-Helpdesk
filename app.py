# =====================================================
# IMPORT LIBRARY
# =====================================================

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    send_from_directory,
    session,
    flash,
    send_file
)

from io import BytesIO
from datetime import datetime
from datetime import timedelta
from zoneinfo import ZoneInfo
import os
import uuid


# =====================================================
# TIMEZONE WIB
# =====================================================

WIB = ZoneInfo("Asia/Jakarta")


def waktu_wib():
    return datetime.now(WIB).replace(tzinfo=None)


# =====================================================
# DATABASE
# =====================================================

from extensions import db
from config import Config

# =====================================================
# MODEL
# =====================================================

from models.ticket import Ticket
from models.admin import Admin
from models.attendance import Attendance

# =====================================================
# SECURITY
# =====================================================

from werkzeug.security import check_password_hash

# =====================================================
# PDF
# =====================================================

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle
)

from reportlab.lib import colors

# =====================================================
# EXCEL
# =====================================================

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment
)

# =====================================================
# FLASK APP
# =====================================================

app = Flask(__name__)

app.config.from_object(Config)


app.secret_key = app.config["SECRET_KEY"]

print(
    "SECRET KEY:",
    app.secret_key
)


app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=2)

app.config.update(

    SESSION_COOKIE_NAME="tollcare_session",

    SESSION_COOKIE_SECURE=False,

    SESSION_COOKIE_HTTPONLY=True,

    SESSION_COOKIE_SAMESITE="Lax",

    SESSION_REFRESH_EACH_REQUEST=True

)

db.init_app(app)

with app.app_context():
    db.create_all()


# =====================================================
# APP CONFIG
# =====================================================

UPLOAD_FOLDER = app.config["UPLOAD_FOLDER"]

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)


# =====================================================
# HELPER DASHBOARD
# =====================================================

def get_dashboard_stats():

    return {

        "total": Ticket.query.count(),

        "open_count": Ticket.query.filter_by(
            status="Open"
        ).count(),

        "process_count": Ticket.query.filter_by(
            status="Diproses"
        ).count(),

        "done_count": Ticket.query.filter_by(
            status="Selesai"
        ).count()

    }

# =====================================================
# BERANDA MASYARAKAT
# =====================================================
@app.route("/")
def index():

    return render_template(
        "masyarakat/index.html"
    )

# =====================================================
# LOGIN ADMIN
# =====================================================

@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():

    # Jika sudah login, langsung ke dashboard
    if "admin" in session:

        return redirect(
            url_for("admin")
        )

    if request.method == "POST":

        username = request.form.get(
            "username",
            ""
        ).strip()

        password = request.form.get(
            "password",
            ""
        )

        # ==========================
        # VALIDASI FORM KOSONG
        # ==========================

        if not username or not password:

            flash(
                "Username dan Password wajib diisi.",
                "warning"
            )

            return redirect(
                url_for("admin_login")
            )

        # ==========================
        # CARI ADMIN
        # ==========================

        admin = Admin.query.filter_by(
            username=username
        ).first()

        # ==========================
        # CEK ADMIN
        # ==========================

        print("ADMIN DITEMUKAN:", admin)

        # ==========================
        # VALIDASI PASSWORD
        # ==========================

        if admin and check_password_hash(
            admin.password,
            password
        ):

            # ==========================
            # BUAT SESSION
            # ==========================

            session.permanent = True

            session["admin"] = admin.username

            print(
                "LOGIN SESSION:",
                session
            )

            flash(
                f"Selamat datang, {admin.fullname}.",
                "success"
            )

            return redirect(
                url_for("admin")
            )

        print(
            "LOGIN GAGAL"
        )

        flash(
            "Username atau Password salah.",
            "danger"
        )

    return render_template(
        "admin/login.html"
    )

# =====================================================
# LOGOUT
# =====================================================
@app.route("/logout")
def logout():

    session.clear()

    flash(
        "Logout berhasil.",
        "success"
    )

    return redirect(
        url_for("admin_login")
    )

# =====================================================
# TRACKING TIKET
# =====================================================
@app.route("/tracking", methods=["GET"])
def tracking():

    ticket = None

    ticket_number = request.args.get(
        "ticket_number",
        ""
    ).strip()

    if ticket_number:

        ticket = Ticket.query.filter_by(
            ticket_number=ticket_number
        ).first()

    return render_template(
        "masyarakat/tracking.html",
        ticket=ticket,
        ticket_number=ticket_number
    )

# =====================================================
# FORM LAPORAN
# =====================================================
@app.route("/report", methods=["GET", "POST"])
def report():

    if request.method == "GET":
        session.pop("_flashes", None)


    if request.method == "POST":

        print("======================")
        print("POST REPORT MASUK")
        print(request.form)
        print(request.files)
        print("======================")


        # ==========================
        # AMBIL DATA
        # ==========================

        fullname = request.form.get("fullname", "").strip()
        phone = request.form.get("phone", "").strip()
        toll_gate = request.form.get("toll_gate", "").strip()
        category = request.form.get("category", "").strip()
        description = request.form.get("description", "").strip()



        # ==========================
        # VALIDASI
        # ==========================

        if not all([
            fullname,
            phone,
            toll_gate,
            category,
            description
        ]):

            flash(
                "Semua data wajib diisi.",
                "warning"
            )

            return redirect(
                url_for("report")
            )



        filename = ""



        # ==========================
        # UPLOAD FOTO
        # ==========================

        photo = request.files.get("photo")


        if photo and photo.filename:


            print(
                "FOTO:",
                photo.filename
            )

            print(
                "TYPE:",
                photo.content_type
            )



            # cek ukuran

            photo.seek(0)

            size = len(photo.read())

            photo.seek(0)


            if size > 5 * 1024 * 1024:

                flash(
                    "Ukuran foto maksimal 5 MB.",
                    "danger"
                )

                return redirect(
                    url_for("report")
                )



            # ==========================
            # TENTUKAN EXTENSI
            # ==========================

            mime = photo.content_type


            if mime == "image/png":

                ext = ".png"


            elif mime == "image/webp":

                ext = ".webp"


            else:

                # default jpg

                ext = ".jpg"



            filename = (
                uuid.uuid4().hex
                +
                ext
            )



            # pastikan folder ada

            os.makedirs(
                app.config["UPLOAD_FOLDER"],
                exist_ok=True
            )



            upload_path = os.path.join(
                app.config["UPLOAD_FOLDER"],
                filename
            )


            print(
                "SIMPAN:",
                upload_path
            )


            photo.save(
                upload_path
            )



        print(
            "FOTO TERSIMPAN:",
            filename
        )



        # ==========================
        # NOMOR TIKET
        # ==========================

        ticket_number = (

            "JM-"

            +

            waktu_wib().strftime(
                "%Y%m%d%H%M%S"
            )

            +

            "-"

            +

            uuid.uuid4().hex[:4].upper()

        )



        # ==========================
        # SIMPAN DATABASE
        # ==========================

        ticket = Ticket(

            ticket_number=ticket_number,

            fullname=fullname,

            phone=phone,

            toll_gate=toll_gate,

            category=category,

            description=description,

            photo=filename,

            status="Open"

        )


        try:

            db.session.add(ticket)

            db.session.commit()


            print(
                "DATABASE BERHASIL"
            )


        except Exception as e:


            db.session.rollback()


            print(
                "DATABASE ERROR:",
                e
            )


            flash(
                "Gagal menyimpan laporan.",
                "danger"
            )


            return redirect(
                url_for("report")
            )



        flash(
            f"Laporan berhasil dibuat. Nomor Tiket: {ticket_number}",
            "success"
        )


        return redirect(
            url_for(
                "tracking",
                ticket_number=ticket_number
            )
        )



    return render_template(
        "masyarakat/report.html"
    )

# =====================================================
# FOTO UPLOAD
# =====================================================

@app.route("/uploads/<path:filename>")
def uploaded_file(filename):

    return send_from_directory(
        app.config["UPLOAD_FOLDER"],
        filename
    )


# =====================================================
# FOTO PERBAIKAN
# =====================================================

@app.route("/uploads/repair/<filename>")
def repair_file(filename):

    return send_from_directory(
        os.path.join(
            app.config["UPLOAD_FOLDER"],
            "repair"
        ),
        filename
    )
# =====================================================
# DASHBOARD ADMIN
# =====================================================
@app.route("/admin")
def admin():

    # ==========================
    # DEBUG SESSION
    # ==========================
    print(
        "ADMIN PAGE SESSION:",
        session
    )


    # ==========================
    # CEK LOGIN
    # ==========================
    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )



    # ==========================
    # DATA ADMIN
    # ==========================
    admin_data = Admin.query.filter_by(

        username=session["admin"]

    ).first()



    # ==========================
    # CEK DATA ADMIN
    # ==========================
    if not admin_data:

        session.clear()

        flash(
            "Data admin tidak ditemukan.",
            "danger"
        )

        return redirect(
            url_for("admin_login")
        )



    # ==========================
    # PAGINATION
    # ==========================
    page = request.args.get(

        "page",

        default=1,

        type=int

    )


    per_page = 10



    tickets = (

        Ticket.query

        .order_by(

            Ticket.created_at.desc()

        )

        .paginate(

            page=page,

            per_page=per_page,

            error_out=False

        )

    )



    # ==========================
    # STATISTIK
    # ==========================
    stats = get_dashboard_stats()



    # ==========================
    # RENDER DASHBOARD
    # ==========================
    return render_template(

        "admin/dashboard.html",

        admin=admin_data,

        tickets=tickets,

        page=page,

        **stats

    )
# =====================================================
# DATA LAPORAN
# =====================================================
@app.route("/admin/laporan")
def admin_laporan():

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )

    admin = Admin.query.filter_by(
        username=session["admin"]
    ).first()

    page = request.args.get(
        "page",
        default=1,
        type=int
    )

    tickets = (

        Ticket.query

        .order_by(
            Ticket.created_at.desc()
        )

        .paginate(

            page=page,

            per_page=10,

            error_out=False

        )

    )

    return render_template(

        "admin/laporan.html",
      tickets=tickets,

        admin=admin,

  
        page=page

    )

# =====================================================
# STATISTIK
# =====================================================
@app.route("/admin/statistik")
def admin_statistik():

    if "admin" not in session:

        return redirect(
         url_for("admin_login")
        )

    admin = Admin.query.filter_by(

        username=session["admin"]

    ).first()

    stats = get_dashboard_stats()

    return render_template(

        "admin/statistik.html",

        admin=admin,

        **stats

    )
# =====================================================
# ABSENSI ADMIN
# =====================================================

@app.route("/admin/absensi")
def admin_absensi():

    # ==========================
    # CEK LOGIN
    # ==========================

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    # ==========================
    # DATA ADMIN
    # ==========================

    admin = Admin.query.filter_by(
        username=session["admin"]
    ).first()


    if not admin:

        session.clear()

        flash(
            "Data admin tidak ditemukan.",
            "danger"
        )

        return redirect(
            url_for("admin_login")
        )


    # ==========================
    # TANGGAL HARI INI
    # ==========================

    today = waktu_wib().date()


    # ==========================
    # ABSENSI HARI INI
    # ==========================

    attendance = Attendance.query.filter_by(
        admin_id=admin.id,
        attendance_date=today
    ).first()


    # ==========================
    # RIWAYAT ABSENSI
    # ==========================

    attendances = (
        Attendance.query
        .filter_by(admin_id=admin.id)
        .order_by(
            Attendance.attendance_date.desc()
        )
        .limit(30)
        .all()
    )


    # ==========================
    # RENDER
    # ==========================

    return render_template(

        "admin/absensi.html",

        admin=admin,

        attendance=attendance,

        attendances=attendances,

        now=waktu_wib()

    )


# =====================================================
# CHECK IN
# =====================================================

@app.route(
    "/admin/absensi/check-in",
    methods=["POST"]
)
def admin_check_in():

    # ==========================
    # CEK LOGIN
    # ==========================

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    # ==========================
    # DATA ADMIN
    # ==========================

    admin = Admin.query.filter_by(
        username=session["admin"]
    ).first()


    if not admin:

        session.clear()

        flash(
            "Data admin tidak ditemukan.",
            "danger"
        )

        return redirect(
            url_for("admin_login")
        )


    # ==========================
    # WAKTU SEKARANG WIB
    # ==========================

    current_time = waktu_wib()

    today = current_time.date()


    # ==========================
    # CEK ABSENSI HARI INI
    # ==========================

    attendance = Attendance.query.filter_by(
        admin_id=admin.id,
        attendance_date=today
    ).first()


    # ==========================
    # SUDAH CHECK IN
    # ==========================

    if attendance and attendance.check_in:

        flash(
            "Anda sudah melakukan Check In hari ini.",
            "warning"
        )

        return redirect(
            url_for("admin_absensi")
        )


    # ==========================
    # BUAT DATA ABSENSI
    # ==========================

    if not attendance:

        attendance = Attendance(

            admin_id=admin.id,

            attendance_date=today,

            check_in=current_time,

            status="Hadir"

        )

        db.session.add(attendance)

    else:

        attendance.check_in = current_time

        attendance.status = "Hadir"


    # ==========================
    # SIMPAN
    # ==========================

    try:

        db.session.commit()

        flash(
            "Check In berhasil dicatat.",
            "success"
        )

    except Exception as e:

        db.session.rollback()

        print(
            "CHECK IN ERROR:",
            e
        )

        flash(
            "Gagal menyimpan Check In.",
            "danger"
        )


    return redirect(
        url_for("admin_absensi")
    )


# =====================================================
# CHECK OUT
# =====================================================

@app.route(
    "/admin/absensi/check-out",
    methods=["POST"]
)
def admin_check_out():

    # ==========================
    # CEK LOGIN
    # ==========================

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    # ==========================
    # DATA ADMIN
    # ==========================

    admin = Admin.query.filter_by(
        username=session["admin"]
    ).first()


    if not admin:

        session.clear()

        flash(
            "Data admin tidak ditemukan.",
            "danger"
        )

        return redirect(
            url_for("admin_login")
        )


    # ==========================
    # WAKTU SEKARANG
    # ==========================

    current_time = waktu_wib()

    today = current_time.date()


    # ==========================
    # CARI ABSENSI
    # ==========================

    attendance = Attendance.query.filter_by(
        admin_id=admin.id,
        attendance_date=today
    ).first()


    # ==========================
    # BELUM CHECK IN
    # ==========================

    if not attendance or not attendance.check_in:

        flash(
            "Anda harus Check In terlebih dahulu.",
            "warning"
        )

        return redirect(
            url_for("admin_absensi")
        )


    # ==========================
    # SUDAH CHECK OUT
    # ==========================

    if attendance.check_out:

        flash(
            "Anda sudah melakukan Check Out hari ini.",
            "warning"
        )

        return redirect(
            url_for("admin_absensi")
        )


    # ==========================
    # SIMPAN CHECK OUT
    # ==========================

    attendance.check_out = current_time


    try:

        db.session.commit()

        flash(
            "Check Out berhasil dicatat.",
            "success"
        )

    except Exception as e:

        db.session.rollback()

        print(
            "CHECK OUT ERROR:",
            e
        )

        flash(
            "Gagal menyimpan Check Out.",
            "danger"
        )


    return redirect(
        url_for("admin_absensi")
    )

# =====================================================
# PROFIL ADMIN
# =====================================================
@app.route("/admin/profile")
def admin_profile():

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    admin = Admin.query.filter_by(

        username=session["admin"]

    ).first()


    stats = get_dashboard_stats()


    return render_template(

        "admin/profile.html",

        admin=admin,

        **stats

    )



# =====================================================
# HALAMAN EDIT PROFIL ADMIN
# =====================================================
@app.route("/admin/profile/edit")
def edit_profile():

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    admin = Admin.query.filter_by(

        username=session["admin"]

    ).first()


    if not admin:

        flash(
            "Data admin tidak ditemukan.",
            "danger"
        )

        return redirect(
            url_for("admin_login")
        )


    return render_template(

        "admin/edit_profile.html",

        admin=admin

    )



# =====================================================
# UPDATE PROFIL ADMIN
# =====================================================
@app.route(
    "/admin/profile/update",
    methods=["POST"]
)
def update_profile():


    # ==========================
    # CEK LOGIN
    # ==========================

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )



    # ==========================
    # AMBIL DATA ADMIN
    # ==========================

    admin = Admin.query.filter_by(

        username=session["admin"]

    ).first()


    if not admin:

        flash(
            "Data admin tidak ditemukan.",
            "danger"
        )

        return redirect(
            url_for("admin_login")
        )



    # ==========================
    # UPDATE DATA PROFILE
    # ==========================

    fullname = request.form.get(
        "fullname",
        ""
    ).strip()


    email = request.form.get(
        "email",
        ""
    ).strip()


    phone = request.form.get(
        "phone",
        ""
    ).strip()



    if fullname:

        admin.fullname = fullname


    admin.email = email

    admin.phone = phone



    # ==========================
    # UPLOAD FOTO PROFIL
    # ==========================

    photo = request.files.get(
        "photo"
    )



    if photo and photo.filename != "":



        # ==========================
        # CEK UKURAN FOTO
        # ==========================

        if photo.content_length and photo.content_length > 5 * 1024 * 1024:


            flash(
                "Ukuran foto maksimal 5 MB.",
                "danger"
            )


            return redirect(
                url_for("admin_profile")
            )



        # ==========================
        # CEK FORMAT FOTO
        # ==========================

        ext = os.path.splitext(

            photo.filename

        )[1].lower()



        allowed = [

            ".jpg",

            ".jpeg",

            ".png"

        ]



        if ext not in allowed:


            flash(
                "Format foto harus JPG, JPEG, atau PNG.",
                "danger"
            )


            return redirect(
                url_for("admin_profile")
            )



        # ==========================
        # BUAT NAMA FILE BARU
        # ==========================

        filename = (

            uuid.uuid4().hex

            +

            ext

        )



        # ==========================
        # FOLDER FOTO ADMIN
        # ==========================

        upload_folder = os.path.join(

            app.config["UPLOAD_FOLDER"],

            "admin"

        )



        os.makedirs(

            upload_folder,

            exist_ok=True

        )



        # ==========================
        # SIMPAN FOTO
        # ==========================

        photo.save(

            os.path.join(

                upload_folder,

                filename

            )

        )



        # SIMPAN NAMA FOTO KE DATABASE

        admin.photo = filename



    # ==========================
    # SIMPAN DATABASE
    # ==========================

    db.session.commit()



    flash(

        "Profil berhasil diperbarui.",

        "success"

    )



    return redirect(

        url_for("admin_profile")

    )

#==================================================
# FOTO  ===ADMIN
# =====================================================
@app.route("/uploads/admin/<filename>")
def admin_photo(filename):

    return send_from_directory(

        os.path.join(

            app.config["UPLOAD_FOLDER"],

            "admin"

        ),

        filename

    )
# =====================================================
# EXPORT PDF
# =====================================================

@app.route("/admin/export/pdf")
def export_pdf():

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    tickets = (

        Ticket.query

        .order_by(
            Ticket.created_at.desc()
        )

        .all()

    )


    buffer = BytesIO()


    doc = SimpleDocTemplate(

        buffer

    )


    data = [

        [

            "No",

            "Nomor Tiket",

            "Nama Pelapor",

            "Gerbang Tol",

            "Status"

        ]

    ]


    for index, ticket in enumerate(
        tickets,
        start=1
    ):

        data.append(

            [

                str(index),

                ticket.ticket_number,

                ticket.fullname,

                ticket.toll_gate,

                ticket.status

            ]

        )


    table = Table(
        data
    )


    table.setStyle(

        TableStyle(

            [

                (
                    "BACKGROUND",
                    (0,0),
                    (-1,0),
                    colors.HexColor("#2563eb")
                ),

                (
                    "TEXTCOLOR",
                    (0,0),
                    (-1,0),
                    colors.white
                ),

                (
                    "GRID",
                    (0,0),
                    (-1,-1),
                    1,
                    colors.black
                ),

                (
                    "ALIGN",
                    (0,0),
                    (-1,-1),
                    "CENTER"
                ),

                (
                    "FONT",
                    (0,0),
                    (-1,0),
                    "Helvetica-Bold"
                )

            ]

        )

    )


    doc.build(

        [

            table

        ]

    )


    buffer.seek(0)


    return send_file(

        buffer,

        as_attachment=True,

        download_name="laporan_tollcare.pdf",

        mimetype="application/pdf"

    )

# =====================================================
# EXPORT EXCEL
# =====================================================

@app.route("/admin/export/excel")
def export_excel():

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    tickets = (

        Ticket.query

        .order_by(
            Ticket.created_at.desc()
        )

        .all()

    )


    workbook = Workbook()


    sheet = workbook.active


    sheet.title = "Laporan TollCare"


    headers = [

        "No",

        "Nomor Tiket",

        "Nama Pelapor",

        "No HP",

        "Gerbang Tol",

        "Kategori",

        "Deskripsi",

        "Status",

        "Tanggal"

    ]


    header_fill = PatternFill(

        start_color="2563EB",

        end_color="2563EB",

        fill_type="solid"

    )


    header_font = Font(

        bold=True,

        color="FFFFFF"

    )


    for col, header in enumerate(

        headers,

        start=1

    ):

        cell = sheet.cell(

            row=1,

            column=col

        )


        cell.value = header

        cell.fill = header_fill

        cell.font = header_font

        cell.alignment = Alignment(

            horizontal="center"

        )



    for row, ticket in enumerate(

        tickets,

        start=2

    ):


        sheet.cell(row,1).value = row - 1

        sheet.cell(row,2).value = ticket.ticket_number

        sheet.cell(row,3).value = ticket.fullname

        sheet.cell(row,4).value = ticket.phone

        sheet.cell(row,5).value = ticket.toll_gate

        sheet.cell(row,6).value = ticket.category

        sheet.cell(row,7).value = ticket.description

        sheet.cell(row,8).value = ticket.status


        sheet.cell(row,9).value = (

            ticket.created_at.strftime(

                "%d-%m-%Y %H:%M"

            )

        )



    widths = {

        "A":8,

        "B":22,

        "C":25,

        "D":18,

        "E":22,

        "F":18,

        "G":40,

        "H":15,

        "I":22

    }


    for column, width in widths.items():

        sheet.column_dimensions[column].width = width



    buffer = BytesIO()


    workbook.save(

        buffer

    )


    buffer.seek(0)


    return send_file(

        buffer,

        as_attachment=True,

        download_name="laporan_tollcare.xlsx",

        mimetype=

        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

# =====================================================
# DETAIL TIKET
# =====================================================

@app.route("/admin/ticket/<int:id>")
def ticket_detail(id):

    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    ticket = Ticket.query.get_or_404(id)


    admin = Admin.query.filter_by(

        username=session["admin"]

    ).first()


    return render_template(

        "admin/ticket_detail.html",

        ticket=ticket,

        admin=admin

    )
# =====================================================
# SIMPAN HASIL PERBAIKAN
# =====================================================

@app.route(
    "/ticket/complete/<int:id>",
    methods=["POST"]
)
def complete_ticket(id):


    if "admin" not in session:

        return redirect(
            url_for("admin_login")
        )


    ticket = Ticket.query.get_or_404(id)



    # ==========================
    # DATA PERBAIKAN
    # ==========================

    ticket.technician = request.form.get(
        "technician",
        ""
    )


    ticket.repair_note = request.form.get(
        "repair_note",
        ""
    )


    ticket.status = request.form.get(
        "status",
        "Selesai"
    )



    # ==========================
    # UPLOAD FOTO PERBAIKAN
    # ==========================

    repair_photo = request.files.get(
        "repair_photo"
    )


    if repair_photo and repair_photo.filename != "":


        # ==========================
        # BATAS UKURAN FOTO
        # ==========================

        if repair_photo.content_length > 5 * 1024 * 1024:

            flash(
                "Ukuran foto maksimal 5 MB.",
                "danger"
            )

            return redirect(

                url_for(
                    "ticket_detail",
                    id=id
                )

            )



        # ==========================
        # CEK FORMAT FOTO
        # ==========================

        ext = os.path.splitext(
            repair_photo.filename
        )[1].lower()


        allowed = [

            ".jpg",

            ".jpeg",

            ".png"

        ]


        if ext not in allowed:

            flash(
                "Format foto harus JPG, JPEG, atau PNG.",
                "danger"
            )

            return redirect(

                url_for(
                    "ticket_detail",
                    id=id
                )

            )



        # ==========================
        # NAMA FILE
        # ==========================

        filename = (

            uuid.uuid4().hex

            +

            ext

        )



        # ==========================
        # FOLDER REPAIR
        # ==========================

        upload_folder = os.path.join(

            app.config["UPLOAD_FOLDER"],

            "repair"

        )


        os.makedirs(

            upload_folder,

            exist_ok=True

        )



        # ==========================
        # SIMPAN FOTO
        # ==========================

        repair_photo.save(

            os.path.join(

                upload_folder,

                filename

            )

        )


        ticket.repair_photo = filename



    # ==========================
    # WAKTU SELESAI
    # ==========================

    if ticket.status == "Selesai":

        ticket.completed_at = waktu_wib()

    else:

        ticket.completed_at = None



    # ==========================
    # SIMPAN DATABASE
    # ==========================

    db.session.commit()



    flash(

        "Perbaikan berhasil disimpan.",

        "success"

    )



    return redirect(

        url_for(

            "ticket_detail",

            id=id

        )

    )
# =====================================================
# HEALTH CHECK
# =====================================================

@app.route("/health")
def health():

    return {

        "status": "running",

        "application": "TollCare Helpdesk",

        "time": datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

    }

# =====================================================
# ERROR 404
# =====================================================

@app.errorhandler(404)
def page_not_found(error):

    return render_template(
        "404.html"
    ), 404

# =====================================================
# ERROR 500
# =====================================================

@app.errorhandler(500)
def internal_error(error):

    db.session.rollback()

    return render_template(
        "500.html"
    ), 500

# =====================================================
# RUN APPLICATION
# =====================================================

if __name__ == "__main__":

    app.run(

        host="0.0.0.0",

        port=5000,

        debug=False

    )